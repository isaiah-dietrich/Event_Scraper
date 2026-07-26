"""Excel I/O for the event pipeline: the internal master workbook and the
weekly client digest.

Two distinct outputs share the same columns and banded-table styling:

- The **internal master workbook** (events_master.xlsx, created automatically
  on first run) is the permanent, append-only record of every event ever
  seen. It exists purely as a dedupe store + history for the pipeline; no
  client ever opens or edits it. `append_rows` grows it and
  `read_existing_event_keys` reads its dedupe keys back.
- The **weekly digest** is a fresh, standalone workbook rebuilt from scratch
  every run and emailed to the client. `write_weekly_digest` produces it.

Output rows are written into a self-expanding Excel Table (blue/white
banded style) so each sheet stays sortable/filterable as it grows.
"""

import os

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo

TABLE_STYLE = "TableStyleMedium2"  # Built-in blue header, white/light-blue banded rows.

EVENTS_SHEET_NAME = "Events"
REJECTED_SHEET_NAME = "Rejected Events"

# Statuses whose rows participate in cross-run dedupe (see _event_key,
# _existing_event_keys, _append_to_sheet).
_DEDUPE_STATUSES = {"ok"}

# Internal row keys, in the order they should appear in each sheet.
OUTPUT_COLUMNS = [
    "title",
    "fit_score",
    "confidence",
    "date",
    "scraped_at",
    "start_time",
    "location",
    "signup_link",
    "status",
    "short_description",
    "fit_reason",
]

# Friendly column headers shown in the sheet.
_DISPLAY_HEADERS = {
    "title": "Event Title",
    "fit_score": "Fit Score",
    "confidence": "Confidence",
    "date": "Event Date",
    "scraped_at": "Date Scraped",
    "start_time": "Start Time",
    "location": "Location",
    "signup_link": "Signup Link",
    "status": "Status",
    "short_description": "Description",
    "fit_reason": "Fit Reason",
}
OUTPUT_HEADERS = [_DISPLAY_HEADERS[column] for column in OUTPUT_COLUMNS]

# These two columns hold long free-text and are exempt from column autosizing.
_NO_AUTOSIZE_COLUMNS = {"short_description", "fit_reason"}
_WRAP_COLUMN_WIDTH = 60
_MAX_AUTOSIZE_WIDTH = 50
_AUTOSIZE_PADDING = 2

# A parsed event's "date" is a real datetime.datetime (always midnight; see
# cli.batch._filter_past_events), so it sorts/filters correctly in Excel -
# but openpyxl's default number format for datetimes includes a time
# component, which would show a distracting "00:00:00" on every row. This
# formats it as a calendar date instead. Applying it to every row in the
# column is safe even for the text values kept for blank/unparseable
# dates - Excel only applies a number format to actual numeric/date values,
# so plain text cells just ignore it and display unchanged.
_DATE_COLUMN_NUMBER_FORMAT = "mmmm d, yyyy"


def _is_rejected(row: dict) -> bool:
    """True for a scored event the model is confident is a poor fit (score 1)."""
    return (
        row.get("status") == "ok"
        and row.get("fit_score") == 1
        and str(row.get("confidence", "")).strip().lower() == "high"
    )


def event_key(title: str, date) -> tuple:
    """Builds a dedupe key for an event: title + date.

    Public so cli.batch can build an identical key for an event that hasn't
    been written yet, to skip re-scoring it (see read_existing_event_keys).
    `date` is expected to be a real datetime.datetime for a successfully
    parsed date (see cli.batch._filter_past_events) - openpyxl always reads
    date cells back as datetime.datetime too, which is exactly why a key
    built here from a freshly-filtered event compares equal to the same
    event's key read back from a previous run's output file.

    Deliberately excludes source_url (there is no longer a URL column in
    the output - see OUTPUT_COLUMNS): the same event cross-posted on two
    different sites now dedupes to a single row instead of one per site.
    """
    return (title, date)


def _event_key(row: dict) -> tuple:
    """Builds a dedupe key for an event row: title + date."""
    return event_key(row.get("title", ""), row.get("date", ""))


def _existing_event_keys(sheet) -> set:
    """Collects dedupe keys for every dedupe-eligible row already in the
    sheet (see _DEDUPE_STATUSES)."""
    status_index = OUTPUT_COLUMNS.index("status")
    title_index = OUTPUT_COLUMNS.index("title")
    date_index = OUTPUT_COLUMNS.index("date")

    keys = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[status_index] not in _DEDUPE_STATUSES:
            continue
        keys.add((row[title_index] or "", row[date_index] or ""))
    return keys


def _apply_table(sheet, last_row: int) -> None:
    """Creates or resizes this sheet's table to cover all current rows.

    The table name is derived from the sheet title (fine for the fixed
    "Events"/"Rejected Events"/"New Events" titles this is ever called on).

    Does nothing if the sheet has no data rows yet (last_row < 2, i.e. only
    the header). A table ref spanning just the header row (e.g. "A1:K1") is
    structurally invalid per the Excel table spec, which requires at least
    one row below the header - writing one anyway doesn't error in openpyxl,
    but Excel flags the file as corrupt on open and silently strips the
    Table/AutoFilter out during its repair. This routinely happened on the
    "Rejected Events" sheet whenever a run produced zero rejected rows. The
    table gets created on a later run once the sheet has its first real row.
    """
    if last_row < 2:
        return
    table_name = sheet.title.replace(" ", "") + "Table"
    last_column_letter = get_column_letter(len(OUTPUT_COLUMNS))
    table_ref = f"A1:{last_column_letter}{last_row}"

    if table_name in sheet.tables:
        sheet.tables[table_name].ref = table_ref
    else:
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE, showRowStripes=True, showFirstColumn=False
        )
        sheet.add_table(table)


def _autosize_columns(sheet, last_row: int) -> None:
    """Sizes each column to fit its longest value, except free-text columns."""
    for index, column in enumerate(OUTPUT_COLUMNS, start=1):
        letter = get_column_letter(index)
        if column in _NO_AUTOSIZE_COLUMNS:
            sheet.column_dimensions[letter].width = _WRAP_COLUMN_WIDTH
            for row_number in range(2, last_row + 1):
                sheet.cell(row=row_number, column=index).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            continue

        max_length = len(str(_DISPLAY_HEADERS[column]))
        for row_number in range(2, last_row + 1):
            value = sheet.cell(row=row_number, column=index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[letter].width = min(
            max_length + _AUTOSIZE_PADDING, _MAX_AUTOSIZE_WIDTH
        )


def _apply_date_number_format(sheet, last_row: int) -> None:
    """Formats the "date" column as a calendar date, not datetime-with-time."""
    date_column_index = OUTPUT_COLUMNS.index("date") + 1
    for row_number in range(2, last_row + 1):
        sheet.cell(row=row_number, column=date_column_index).number_format = (
            _DATE_COLUMN_NUMBER_FORMAT
        )


def _validate_header(sheet) -> None:
    """Raises if an existing sheet's header row doesn't match OUTPUT_HEADERS.

    Appending rows in OUTPUT_COLUMNS order to a sheet with a different/older
    header silently misaligns every column after the divergence point (this
    has happened before, when the "date" column was added after some output
    files already existed) - so we fail loudly instead of writing bad data.
    """
    header = [cell.value for cell in sheet[1]]
    if header != OUTPUT_HEADERS:
        raise RuntimeError(
            f"Sheet {sheet.title!r} has header {header}, which does not "
            f"match the current expected columns {OUTPUT_HEADERS}. "
            "Appending would misalign data. Regenerate this output file "
            "(or migrate its header row to match) before running again."
        )


def _get_or_create_sheet(workbook, title: str):
    """Returns the named sheet, creating it with a header row if missing."""
    if title in workbook.sheetnames:
        sheet = workbook[title]
        _validate_header(sheet)
        return sheet
    sheet = workbook.create_sheet(title)
    sheet.append(OUTPUT_HEADERS)
    return sheet


def _append_to_sheet(sheet, rows: list[dict], existing_keys: set) -> int:
    """Appends rows to one sheet, skipping dedupe-eligible rows (see
    _DEDUPE_STATUSES) already in existing_keys.

    Mutates existing_keys with the key of every dedupe-eligible row actually
    written, so a single key set can be shared across sheets to prevent the
    same event ending up duplicated across Events and Rejected Events.
    """
    skipped_count = 0
    for row in rows:
        if row.get("status") in _DEDUPE_STATUSES:
            key = _event_key(row)
            if key in existing_keys:
                skipped_count += 1
                print(f"  [dedup skip] {key[0]} | {key[1]}")
                continue
            existing_keys.add(key)
        sheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])

    _apply_table(sheet, sheet.max_row)
    _autosize_columns(sheet, sheet.max_row)
    _apply_date_number_format(sheet, sheet.max_row)
    return skipped_count


def append_rows(path: str, rows: list[dict]) -> None:
    """Appends result rows to the internal master workbook.

    The master (events_master.xlsx, created here on first run) is the
    permanent, append-only seen-events record the pipeline dedupes against
    across runs - no client ever opens or edits it. Creates the workbook
    with "Events" and "Rejected Events" sheets if they do not already exist.
    A scored event is routed to Rejected Events if the model gave it
    fit_score 1 with high confidence, and everything else (including
    non-"ok" failure/no_events rows) goes to Events. Dedupe-eligible rows
    (see _DEDUPE_STATUSES) are skipped if a row with the same title and date
    is already present in either sheet, so re-running the batch over the same
    sites does not duplicate events or let one flip between sheets across
    runs - and the same event cross-posted on two different sites now
    collapses to a single row (see event_key).

    Args:
        path: Path to the master .xlsx file. Row dicts may still carry a
            "source_url" key (e.g. for building a signup_link fallback
            before this call) - it is simply not written as a column.
        rows: A list of dicts, each keyed by a subset of OUTPUT_COLUMNS plus
            any extra fields used only for dedupe (e.g. the event's "date").
    """
    if os.path.exists(path):
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    events_sheet = _get_or_create_sheet(workbook, EVENTS_SHEET_NAME)
    rejected_sheet = _get_or_create_sheet(workbook, REJECTED_SHEET_NAME)

    existing_keys = _existing_event_keys(events_sheet) | _existing_event_keys(rejected_sheet)
    normal_rows = [row for row in rows if not _is_rejected(row)]
    rejected_rows = [row for row in rows if _is_rejected(row)]

    skipped_count = _append_to_sheet(events_sheet, normal_rows, existing_keys)
    skipped_count += _append_to_sheet(rejected_sheet, rejected_rows, existing_keys)

    workbook.save(path)
    if skipped_count:
        print(f"Skipped {skipped_count} duplicate event(s) already in {path}")


def read_existing_event_keys(path: str) -> set:
    """Returns the dedupe keys of every event already in the master workbook.

    This is the union of _existing_event_keys over both the Events and
    Rejected Events sheets, i.e. exactly the set append_rows would dedupe
    a new run's rows against. Callers (see cli.batch) use it to skip
    scoring an event before it's even written, rather than paying for a
    scoring AI call only to have the row silently dropped by write-time
    dedupe in append_rows.

    Returns an empty set if no file exists at `path` yet - the natural
    "nothing known" starting point for a first run.
    """
    if not os.path.exists(path):
        return set()
    workbook = load_workbook(path)
    keys = set()
    for title in (EVENTS_SHEET_NAME, REJECTED_SHEET_NAME):
        if title not in workbook.sheetnames:
            continue
        sheet = workbook[title]
        _validate_header(sheet)
        keys |= _existing_event_keys(sheet)
    return keys


# --- Weekly client digest ----------------------------------------------------
#
# write_weekly_digest builds the fresh, standalone workbook emailed to the
# client each week. Unlike the master, it is rebuilt from scratch every run
# and carries no dedupe/history state - the caller has already deduped this
# run's rows against the master before handing them here.


def write_weekly_digest(path: str, rows: list[dict]) -> None:
    """Writes a fresh weekly digest workbook the client receives by email.

    A brand-new workbook is created at `path` every call (any existing file
    is overwritten) - each week's digest is a standalone snapshot, not an
    accumulating record. Only rows whose status is "ok" are included;
    failed/no_events rows are reported in the email body by other code and
    never appear here. The "ok" rows are split with _is_rejected into a
    "New Events" sheet (good fits) and a "Rejected Events" sheet (confident
    poor fits); both sheets are always created with a header row even when
    empty.

    Rows are written in OUTPUT_COLUMNS order with no dedupe - the caller has
    already deduped against the master. The signup_link "*" convention
    (base-URL fallback flagged with a trailing " *" when extraction found no
    direct link) passes through untouched, since rows already carry it. The
    digest deliberately reuses the same columns and banded-table styling as
    the master workbook so the client sees a familiar format.

    Args:
        path: Path to write the digest .xlsx to (overwritten if it exists).
        rows: This run's result dicts, keyed by a subset of OUTPUT_COLUMNS
            plus "status". Non-"ok" rows are ignored.
    """
    ok_rows = [row for row in rows if row.get("status") == "ok"]
    new_rows = [row for row in ok_rows if not _is_rejected(row)]
    rejected_rows = [row for row in ok_rows if _is_rejected(row)]

    workbook = Workbook()
    workbook.remove(workbook.active)

    for title, sheet_rows in (
        ("New Events", new_rows),
        (REJECTED_SHEET_NAME, rejected_rows),
    ):
        sheet_rows = sorted(sheet_rows, key=lambda row: row.get("fit_score", 0), reverse=True)
        sheet = workbook.create_sheet(title)
        sheet.append(OUTPUT_HEADERS)
        for row in sheet_rows:
            sheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])
        # _apply_table no-ops on a header-only sheet (a header-only table ref
        # corrupts the file), so an empty digest sheet is left table-less.
        _apply_table(sheet, sheet.max_row)
        _autosize_columns(sheet, sheet.max_row)
        _apply_date_number_format(sheet, sheet.max_row)

    workbook.save(path)
