"""
Batch event discovery: read a list of site URLs from an input spreadsheet,
run the FETCH -> REDUCE -> EXTRACT pipeline against each one, and append the
results to a master output spreadsheet.

Scoring is intentionally skipped for now (see scrape_events.score_event for
the still-available scoring step, to be wired back in later).
"""

import datetime
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from anthropic import Anthropic

from scrape_events import fetch_rendered_html, reduce_html, extract_events, EXTRACTION_FIELDS

INPUT_PATH = "websites.xlsx"
OUTPUT_PATH = "events_output.xlsx"

# Each site gets its own browser instance and its own LLM call, run
# concurrently. Each worker pops a visible Chromium window (headless=False,
# see scrape_events.fetch_rendered_html), so don't set this too high locally.
MAX_WORKERS = 4

OUTPUT_HEADERS = [
    "scraped_at",
    "source_url",
    "status",
    *EXTRACTION_FIELDS,
]


def read_input_urls(path: str) -> list:
    """Read site URLs from the input spreadsheet's first column (header row skipped)."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Input spreadsheet not found: {path}")
    wb = load_workbook(path)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            urls.append(str(row[0]).strip())
    return urls


def append_rows(path: str, rows: list):
    """Append result rows to the master output spreadsheet, creating it if needed."""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Events"
        ws.append(OUTPUT_HEADERS)

    for row in rows:
        ws.append([row.get(col, "") for col in OUTPUT_HEADERS])

    wb.save(path)


def process_site(client: Anthropic, url: str) -> list:
    """Run FETCH -> REDUCE -> EXTRACT for one site. Returns one or more output rows."""
    timestamp = datetime.datetime.now().isoformat(timespec="seconds")
    base = {"scraped_at": timestamp, "source_url": url}

    try:
        html = fetch_rendered_html(url)
    except RuntimeError as e:
        return [{**base, "status": f"failed: {e}"}]

    page_text = reduce_html(html)
    if not page_text:
        return [{**base, "status": "failed: empty page text after reduction"}]

    try:
        events = extract_events(client, page_text)
    except ValueError as e:
        return [{**base, "status": f"failed: {e}"}]

    if not events:
        return [{**base, "status": "no_events"}]

    rows = []
    for event in events:
        row = {**base, "status": "ok"}
        row.update({field: event.get(field, "") for field in EXTRACTION_FIELDS})
        rows.append(row)
    return rows


def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY environment variable is not set.", file=sys.stderr)
        sys.exit(1)
    client = Anthropic(api_key=api_key)

    try:
        urls = read_input_urls(INPUT_PATH)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    if not urls:
        print(f"No URLs found in {INPUT_PATH}. Nothing to do.")
        sys.exit(0)

    print(f"Processing {len(urls)} site(s) from {INPUT_PATH} (up to {MAX_WORKERS} at a time)...")
    all_rows = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_url = {executor.submit(process_site, client, url): url for url in urls}
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                rows = future.result()
            except Exception as e:
                rows = [{
                    "scraped_at": datetime.datetime.now().isoformat(timespec="seconds"),
                    "source_url": url,
                    "status": f"failed: unexpected error: {e}",
                }]
            print(f"[done] {url}")
            for row in rows:
                print(f"    -> {row['status']}" + (f": {row.get('title')}" if row["status"] == "ok" else ""))
            all_rows.extend(rows)

    append_rows(OUTPUT_PATH, all_rows)
    print(f"\nAppended {len(all_rows)} row(s) to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
