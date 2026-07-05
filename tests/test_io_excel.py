import datetime

import pytest
from openpyxl import load_workbook
from openpyxl import Workbook

from utility.io_excel import _is_rejected
from utility.io_excel import _sheet_title_from_url
from utility.io_excel import _unique_sheet_title
from utility.io_excel import _unique_table_name
from utility.io_excel import append_rows
from utility.io_excel import EVENTS_SHEET_NAME
from utility.io_excel import OUTPUT_COLUMNS
from utility.io_excel import OUTPUT_HEADERS
from utility.io_excel import read_input_urls
from utility.io_excel import REJECTED_SHEET_NAME
from utility.io_excel import write_per_site_sheets


def _ok_row(**overrides):
    row = {
        "title": "AI Meetup",
        "fit_score": 4,
        "confidence": "high",
        "date": "July 4, 2026",
        "scraped_at": "June 30, 2026",
        "source_url": "https://example.com",
        "start_time": "6:00 PM",
        "location": "Atlanta, GA",
        "status": "ok",
        "short_description": "A meetup about AI.",
        "fit_reason": "Great fit.",
    }
    row.update(overrides)
    return row


# --- read_input_urls -------------------------------------------------------


def test_read_input_urls_returns_non_empty_stripped_urls(tmp_path):
    path = tmp_path / "sites.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["url"])
    sheet.append(["https://a.example.com "])
    sheet.append([None])
    sheet.append(["  https://b.example.com"])
    sheet.append([""])
    workbook.save(path)

    urls = read_input_urls(str(path))

    assert urls == ["https://a.example.com", "https://b.example.com"]


def test_read_input_urls_missing_file_raises(tmp_path):
    missing = tmp_path / "does_not_exist.xlsx"

    with pytest.raises(FileNotFoundError):
        read_input_urls(str(missing))


# --- _is_rejected ------------------------------------------------------


def test_is_rejected_true_for_high_confidence_score_one():
    assert _is_rejected(_ok_row(fit_score=1, confidence="high")) is True


def test_is_rejected_false_for_low_confidence_score_one():
    assert _is_rejected(_ok_row(fit_score=1, confidence="low")) is False


def test_is_rejected_false_for_score_above_one():
    assert _is_rejected(_ok_row(fit_score=2, confidence="high")) is False


def test_is_rejected_false_for_non_ok_status():
    assert _is_rejected(_ok_row(fit_score=1, confidence="high", status="no_events")) is False


# --- append_rows: sheet creation and routing --------------------------------


def test_append_rows_creates_workbook_with_expected_headers(tmp_path):
    path = tmp_path / "out.xlsx"

    append_rows(str(path), [_ok_row()])

    workbook = load_workbook(path)
    assert set(workbook.sheetnames) >= {EVENTS_SHEET_NAME, REJECTED_SHEET_NAME}
    events_header = [cell.value for cell in workbook[EVENTS_SHEET_NAME][1]]
    assert events_header == OUTPUT_HEADERS


def test_append_rows_routes_good_score_to_events_sheet(tmp_path):
    path = tmp_path / "out.xlsx"

    append_rows(str(path), [_ok_row(title="Good Event", fit_score=4, confidence="high")])

    workbook = load_workbook(path)
    events_titles = [row[0] for row in workbook[EVENTS_SHEET_NAME].iter_rows(min_row=2, values_only=True)]
    rejected_titles = [row[0] for row in workbook[REJECTED_SHEET_NAME].iter_rows(min_row=2, values_only=True)]
    assert "Good Event" in events_titles
    assert "Good Event" not in rejected_titles


def test_append_rows_routes_rejected_score_to_rejected_sheet(tmp_path):
    path = tmp_path / "out.xlsx"

    append_rows(str(path), [_ok_row(title="Bad Event", fit_score=1, confidence="high")])

    workbook = load_workbook(path)
    events_titles = [row[0] for row in workbook[EVENTS_SHEET_NAME].iter_rows(min_row=2, values_only=True)]
    rejected_titles = [row[0] for row in workbook[REJECTED_SHEET_NAME].iter_rows(min_row=2, values_only=True)]
    assert "Bad Event" in rejected_titles
    assert "Bad Event" not in events_titles


def test_append_rows_non_ok_status_goes_to_events_sheet(tmp_path):
    path = tmp_path / "out.xlsx"
    failure_row = {
        "scraped_at": "June 30, 2026",
        "source_url": "https://example.com",
        "status": "failed: timeout",
    }

    append_rows(str(path), [failure_row])

    workbook = load_workbook(path)
    events_rows = list(workbook[EVENTS_SHEET_NAME].iter_rows(min_row=2, values_only=True))
    rejected_rows = list(workbook[REJECTED_SHEET_NAME].iter_rows(min_row=2, values_only=True))
    assert len(events_rows) == 1
    assert len(rejected_rows) == 0


# --- append_rows: dedupe -----------------------------------------------


def test_append_rows_dedupes_identical_ok_event_across_calls(tmp_path):
    path = tmp_path / "out.xlsx"
    row = _ok_row(title="Repeat Event", date="July 4, 2026", source_url="https://x.com")

    append_rows(str(path), [row])
    append_rows(str(path), [row])

    workbook = load_workbook(path)
    events_rows = list(workbook[EVENTS_SHEET_NAME].iter_rows(min_row=2, values_only=True))
    assert len(events_rows) == 1


def test_append_rows_dedupes_across_events_and_rejected_sheets(tmp_path):
    path = tmp_path / "out.xlsx"
    good_row = _ok_row(
        title="Flip Flop Event", date="July 4, 2026", source_url="https://x.com",
        fit_score=4, confidence="high",
    )
    bad_row = _ok_row(
        title="Flip Flop Event", date="July 4, 2026", source_url="https://x.com",
        fit_score=1, confidence="high",
    )

    append_rows(str(path), [good_row])
    append_rows(str(path), [bad_row])

    workbook = load_workbook(path)
    events_rows = list(workbook[EVENTS_SHEET_NAME].iter_rows(min_row=2, values_only=True))
    rejected_rows = list(workbook[REJECTED_SHEET_NAME].iter_rows(min_row=2, values_only=True))
    assert len(events_rows) == 1
    assert len(rejected_rows) == 0


def test_append_rows_does_not_dedupe_different_dates(tmp_path):
    path = tmp_path / "out.xlsx"
    row_a = _ok_row(title="Recurring Event", date="July 4, 2026", source_url="https://x.com")
    row_b = _ok_row(title="Recurring Event", date="August 4, 2026", source_url="https://x.com")

    append_rows(str(path), [row_a])
    append_rows(str(path), [row_b])

    workbook = load_workbook(path)
    events_rows = list(workbook[EVENTS_SHEET_NAME].iter_rows(min_row=2, values_only=True))
    assert len(events_rows) == 2


# --- append_rows: header validation --------------------------------------


def test_append_rows_raises_on_mismatched_existing_header(tmp_path):
    path = tmp_path / "out.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet(EVENTS_SHEET_NAME)
    sheet.append(["Some", "Old", "Header"])
    workbook.create_sheet(REJECTED_SHEET_NAME).append(["Some", "Old", "Header"])
    workbook.save(path)

    with pytest.raises(RuntimeError, match="does not match"):
        append_rows(str(path), [_ok_row()])


# --- append_rows: table grows across calls --------------------------------


def test_append_rows_creates_no_table_for_a_sheet_with_zero_data_rows(tmp_path):
    # A table ref spanning only the header row (e.g. "A1:K1") is invalid per
    # the Excel table spec and gets stripped out with a repair warning on
    # open - this run's only row is "ok", so Rejected Events ends up with
    # zero data rows and must not get a table at all.
    path = tmp_path / "out.xlsx"

    append_rows(str(path), [_ok_row(fit_score=4, confidence="high")])

    workbook = load_workbook(path)
    assert dict(workbook[REJECTED_SHEET_NAME].tables) == {}
    assert len(dict(workbook[EVENTS_SHEET_NAME].tables)) == 1


def test_append_rows_creates_table_once_a_sheet_gets_its_first_row(tmp_path):
    # First call leaves Rejected Events empty (no table yet); second call
    # adds its first rejected row and should create the table then.
    path = tmp_path / "out.xlsx"
    append_rows(str(path), [_ok_row(title="Good Event", fit_score=4, confidence="high")])

    append_rows(str(path), [_ok_row(title="Bad Event", fit_score=1, confidence="high")])

    workbook = load_workbook(path)
    rejected_tables = dict(workbook[REJECTED_SHEET_NAME].tables)
    assert len(rejected_tables) == 1
    table_name = next(iter(rejected_tables))
    assert rejected_tables[table_name].ref == "A1:K2"


def test_append_rows_table_ref_grows_across_calls(tmp_path):
    path = tmp_path / "out.xlsx"
    append_rows(str(path), [_ok_row(title="One", date="July 1, 2026")])
    workbook = load_workbook(path)
    events_sheet = workbook[EVENTS_SHEET_NAME]
    table_name = next(iter(events_sheet.tables))
    first_ref = events_sheet.tables[table_name].ref

    append_rows(str(path), [_ok_row(title="Two", date="July 2, 2026")])
    workbook = load_workbook(path)
    events_sheet = workbook[EVENTS_SHEET_NAME]
    second_ref = events_sheet.tables[table_name].ref

    assert first_ref == "A1:K2"
    assert second_ref == "A1:K3"


# --- append_rows: real date values ---------------------------------------


def test_append_rows_stores_real_datetime_as_an_actual_date_cell(tmp_path):
    path = tmp_path / "out.xlsx"
    row = _ok_row(title="Dated Event", date=datetime.datetime(2026, 10, 12))

    append_rows(str(path), [row])

    workbook = load_workbook(path)
    date_column_index = OUTPUT_COLUMNS.index("date") + 1
    cell = workbook[EVENTS_SHEET_NAME].cell(row=2, column=date_column_index)
    assert cell.value == datetime.datetime(2026, 10, 12)
    assert cell.number_format == "mmmm d, yyyy"


def test_append_rows_dedupes_real_datetime_dates_across_reload(tmp_path):
    # A real date value written in one append_rows call must still compare
    # equal to a freshly-built datetime.datetime for the same day in a
    # later call, even after a save/reload round-trip through openpyxl.
    path = tmp_path / "out.xlsx"
    row = _ok_row(
        title="Repeat Dated Event",
        date=datetime.datetime(2026, 10, 12),
        source_url="https://x.com",
    )

    append_rows(str(path), [row])
    append_rows(str(path), [row])

    workbook = load_workbook(path)
    events_rows = list(workbook[EVENTS_SHEET_NAME].iter_rows(min_row=2, values_only=True))
    assert len(events_rows) == 1


def test_append_rows_keeps_text_date_for_unparseable_dates(tmp_path):
    # Blank/unparseable dates are kept as their original raw text (see
    # cli.batch._filter_past_events) rather than forced into a date value.
    path = tmp_path / "out.xlsx"
    row = _ok_row(title="Weird Date Event", date="sometime next quarter")

    append_rows(str(path), [row])

    workbook = load_workbook(path)
    date_column_index = OUTPUT_COLUMNS.index("date") + 1
    cell = workbook[EVENTS_SHEET_NAME].cell(row=2, column=date_column_index)
    assert cell.value == "sometime next quarter"


# --- _sheet_title_from_url ---------------------------------------------


def test_sheet_title_strips_scheme_and_replaces_forbidden_chars():
    assert _sheet_title_from_url("https://example.com/events") == "example.com-events"
    assert _sheet_title_from_url("http://example.com") == "example.com"


def test_sheet_title_falls_back_to_site_for_empty_result():
    assert _sheet_title_from_url("https://") == "site"


# --- _unique_sheet_title -------------------------------------------------


def test_unique_sheet_title_returns_base_when_no_collision():
    assert _unique_sheet_title("https://a.example.com", set()) == "a.example.com"


def test_unique_sheet_title_dedupes_on_collision():
    used = {"a.example.com"}
    assert _unique_sheet_title("https://a.example.com", used) == "a.example.com (2)"


def test_unique_sheet_title_dedupes_multiple_collisions_in_order():
    used = {"a.example.com", "a.example.com (2)"}
    assert _unique_sheet_title("https://a.example.com", used) == "a.example.com (3)"


def test_unique_sheet_title_truncates_to_excel_limit():
    long_url = "https://" + "a" * 50 + ".example.com"
    title = _unique_sheet_title(long_url, set())
    assert len(title) == 31


def test_unique_sheet_title_truncated_collision_still_fits_and_dedupes():
    long_url = "https://" + "a" * 50 + ".example.com"
    base_title = _unique_sheet_title(long_url, set())
    deduped = _unique_sheet_title(long_url, {base_title.lower()})
    assert len(deduped) <= 31
    assert deduped != base_title


# --- _unique_table_name --------------------------------------------------


def test_unique_table_name_strips_non_word_characters():
    assert _unique_table_name("a.example.com-events", set()) == "aexamplecomeventsTable"


def test_unique_table_name_prefixes_leading_digit():
    assert _unique_table_name("123events", set()) == "T123eventsTable"


def test_unique_table_name_dedupes_on_collision():
    used = {"aexamplecomTable"}
    assert _unique_table_name("a.example.com", used) == "aexamplecomTable2"


# --- write_per_site_sheets ------------------------------------------------


def test_write_per_site_sheets_creates_one_sheet_per_url(tmp_path):
    path = tmp_path / "by_site.xlsx"
    rows_by_url = {
        "https://a.example.com": [_ok_row(title="A Event", source_url="https://a.example.com")],
        "https://b.example.com": [_ok_row(title="B Event", source_url="https://b.example.com")],
    }

    write_per_site_sheets(str(path), rows_by_url)

    workbook = load_workbook(path)
    assert set(workbook.sheetnames) == {"a.example.com", "b.example.com"}
    a_titles = [row[0] for row in workbook["a.example.com"].iter_rows(min_row=2, values_only=True)]
    b_titles = [row[0] for row in workbook["b.example.com"].iter_rows(min_row=2, values_only=True)]
    assert a_titles == ["A Event"]
    assert b_titles == ["B Event"]


def test_write_per_site_sheets_uses_standard_headers(tmp_path):
    path = tmp_path / "by_site.xlsx"

    write_per_site_sheets(str(path), {"https://a.example.com": [_ok_row()]})

    workbook = load_workbook(path)
    header = [cell.value for cell in workbook["a.example.com"][1]]
    assert header == OUTPUT_HEADERS


def test_write_per_site_sheets_formats_date_column_as_date(tmp_path):
    path = tmp_path / "by_site.xlsx"
    row = _ok_row(date=datetime.datetime(2026, 10, 12))

    write_per_site_sheets(str(path), {"https://a.example.com": [row]})

    workbook = load_workbook(path)
    date_column_index = OUTPUT_COLUMNS.index("date") + 1
    cell = workbook["a.example.com"].cell(row=2, column=date_column_index)
    assert cell.value == datetime.datetime(2026, 10, 12)
    assert cell.number_format == "mmmm d, yyyy"


def test_write_per_site_sheets_does_not_split_by_status_or_score(tmp_path):
    path = tmp_path / "by_site.xlsx"
    rows = [
        _ok_row(title="Good Fit", fit_score=5, confidence="high"),
        _ok_row(title="Bad Fit", fit_score=1, confidence="high"),
        {"scraped_at": "June 30, 2026", "source_url": "https://a.example.com", "status": "no_events"},
    ]

    write_per_site_sheets(str(path), {"https://a.example.com": rows})

    workbook = load_workbook(path)
    # Only one sheet for this URL - good, bad, and no-events rows all together.
    assert workbook.sheetnames == ["a.example.com"]
    data_rows = list(workbook["a.example.com"].iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 3


def test_write_per_site_sheets_dedupes_colliding_sheet_titles(tmp_path):
    path = tmp_path / "by_site.xlsx"
    # These two URLs sanitize to the exact same sheet title ("/" and ":" or
    # "?" both become "-"), so both map to "a.example.com-path-1".
    rows_by_url = {
        "https://a.example.com/path:1": [_ok_row(title="First")],
        "https://a.example.com/path?1": [_ok_row(title="Second")],
    }

    write_per_site_sheets(str(path), rows_by_url)

    workbook = load_workbook(path)
    assert len(workbook.sheetnames) == 2
    assert len(set(workbook.sheetnames)) == 2


def test_write_per_site_sheets_overwrites_previous_run(tmp_path):
    path = tmp_path / "by_site.xlsx"
    write_per_site_sheets(str(path), {"https://old.example.com": [_ok_row(title="Old")]})

    write_per_site_sheets(str(path), {"https://new.example.com": [_ok_row(title="New")]})

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["new.example.com"]


def test_write_per_site_sheets_handles_empty_input(tmp_path):
    path = tmp_path / "by_site.xlsx"

    write_per_site_sheets(str(path), {})

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["No Sites"]
