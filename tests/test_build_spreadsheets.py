from openpyxl import load_workbook

import cli.build_spreadsheets as build_spreadsheets
from utility.io_excel import EVENTS_SHEET_NAME
from utility.io_excel import OUTPUT_HEADERS
from utility.io_excel import REJECTED_SHEET_NAME


def test_build_input_writes_header_and_sample_urls(tmp_path, monkeypatch):
    input_path = tmp_path / "websites.xlsx"
    monkeypatch.setattr(build_spreadsheets, "INPUT_PATH", str(input_path))

    build_spreadsheets.build_input()

    workbook = load_workbook(input_path)
    sheet = workbook.active
    assert sheet.title == "Sites"
    rows = [row[0] for row in sheet.iter_rows(min_row=1, values_only=True)]
    assert rows[0] == "url"
    assert rows[1:] == build_spreadsheets.SAMPLE_SITES


def test_build_output_template_creates_empty_sheets_with_headers(tmp_path, monkeypatch):
    output_path = tmp_path / "events_output.xlsx"
    monkeypatch.setattr(build_spreadsheets, "OUTPUT_PATH", str(output_path))

    build_spreadsheets.build_output_template()

    workbook = load_workbook(output_path)
    assert set(workbook.sheetnames) >= {EVENTS_SHEET_NAME, REJECTED_SHEET_NAME}
    events_header = [cell.value for cell in workbook[EVENTS_SHEET_NAME][1]]
    assert events_header == OUTPUT_HEADERS
    assert workbook[EVENTS_SHEET_NAME].max_row == 1  # header row only, no data rows


def test_build_output_template_overwrites_existing_file(tmp_path, monkeypatch):
    output_path = tmp_path / "events_output.xlsx"
    output_path.write_text("stale, not a real workbook")
    monkeypatch.setattr(build_spreadsheets, "OUTPUT_PATH", str(output_path))

    build_spreadsheets.build_output_template()

    # Loads cleanly as a real workbook now, proving the stale file was replaced.
    workbook = load_workbook(output_path)
    assert EVENTS_SHEET_NAME in workbook.sheetnames
